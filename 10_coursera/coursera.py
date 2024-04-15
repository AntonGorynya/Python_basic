import requests
from bs4 import BeautifulSoup
import random
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import argparse
import re


XML_COURSE_LIST = 'https://www.coursera.org/sitemap~www~courses.xml'
COURSE_COUNT = 1


def get_courses_list():
    response = requests.get(XML_COURSE_LIST)
    soup = BeautifulSoup(response.text, 'lxml')
    course_list = [course.text for course in soup.body.find_all('loc')]
    return course_list


def get_course_page(course_url):
    response = requests.get(url=course_url)
    response.encoding = 'UTF-8'
    html = BeautifulSoup(response.text, 'html.parser')
    return html


def get_course_info(html):
    course_name = html.find('h1', class_='title display-3-text')
    if course_name:
        course_name = course_name.text
    else:
        course_name = None
    course_lang = html.find('div', class_='rc-Language')
    if course_lang:
        course_lang = course_lang.text
    else:
        course_lang = None
    assesment = html.find('div',
                          class_='rc-RatingsHeader horizontal-box'
                                 ' align-items-absolute-center')
    if assesment:
        assesment = re.search(r'[\d.]+', assesment.text).group()
    else:
        assesment = None
    start_date = html.find('div',
                           class_='startdate rc-StartDateString'
                                  ' caption-text')
    if start_date:
        start_date = re.search(r'\w+ \d+', start_date.text).group()
    else:
        start_date = None
    course_duration = html.find_all('div',
                                    class_='week-heading body-2-text')
    if course_duration:
        course_duration = course_duration[-1].text[5:]
    else:
        course_duration = None
    return {'course_name': course_name,
            'course_lang': course_lang,
            'assesment': assesment,
            'start_date': start_date,
            'course_duration': course_duration}


def create_table_content(courses_info):
    table_head_param = [{'Course name': 30.0},
                        {'Course lang': 10.0},
                        {'Assesment': 10.0},
                        {'Start date': 10.0},
                        {'Course duration in weeks': 20.0}]
    table_content = [[course_info['course_name'],
                      course_info['course_lang'],
                      course_info['assesment'],
                      course_info['start_date'],
                      course_info['course_duration']]
                     for course_info in courses_info]
    return table_head_param, table_content


def output_courses_info_to_xlsx(table_head_param, table_content):
    wb = Workbook()
    ws1 = wb.active
    table_head = []
    for col_number, col_param in enumerate(table_head_param):
        col_name, col_width = list(col_param.items())[0]
        letter = get_column_letter(col_number+1)
        ws1.column_dimensions[letter].width = col_width
        table_head.append(col_name)
    ws1.append(table_head)
    for exel_row in table_content:
        ws1.append(exel_row)
    return wb


def save_workbook(workbook, filepath):
    workbook.save(filepath)


def create_parser():
    parser = argparse.ArgumentParser(description='course_info')
    parser.add_argument("output_file", nargs='?', const=1,
                        default='book.xlsx',
                        type=str, help="path to output file")
    return parser


if __name__ == '__main__':
    parser = create_parser()
    args = parser.parse_args()
    courses_info = [get_course_info(get_course_page(course_url))
                    for course_url in random.sample(get_courses_list(),
                                                    COURSE_COUNT)]
    workbook = output_courses_info_to_xlsx(*create_table_content(courses_info))
    save_workbook(workbook, args.output_file)
